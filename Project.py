from tkinter import * 
from tkinter import messagebox, filedialog
import fitz as fi
import openpyxl
from openpyxl.styles import Alignment
import pytesseract
from pdf2image import convert_from_path
import cv2
import re
import numpy as np
import os
import time
from io import BytesIO
from azure.cognitiveservices.vision.computervision import ComputerVisionClient
from azure.cognitiveservices.vision.computervision.models import OperationStatusCodes
from msrest.authentication import CognitiveServicesCredentials
import re
from spellchecker import SpellChecker
from io import BytesIO
from azure.cognitiveservices.vision.computervision import ComputerVisionClient
from azure.cognitiveservices.vision.computervision.models import OperationStatusCodes
from msrest.authentication import CognitiveServicesCredentials

def rotate_image(image, angle):
    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    matrix = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated_image = cv2.warpAffine(image, matrix, (w, h), flags=cv2.INTER_CUBIC)
    return rotated_image

def handwritten(images):
    subscription_key = 'subscription key'
    endpoint = 'endpoint'
    
    computervision_client = ComputerVisionClient(endpoint, CognitiveServicesCredentials(subscription_key))
    texts = []

    for image in images:
        open_cv_image = np.array(image)
        _, buffer = cv2.imencode('.png', open_cv_image)
        image_stream = BytesIO(buffer)

        ocr_result = computervision_client.read_in_stream(image_stream, raw=True)
        operation_location = ocr_result.headers['Operation-Location']
        operation_id = operation_location.split('/')[-1]

        while True:
            result = computervision_client.get_read_result(operation_id)
            if result.status not in ['notStarted', 'running']:
                break
            time.sleep(1)

        if result.status == OperationStatusCodes.succeeded:
            for page in result.analyze_result.read_results:
                for line in page.lines:
                    filtered_words = [word for word in line.text.split() if len(word) > 5]
                    if filtered_words:
                        texts.append(' '.join(filtered_words))

    return texts

def iconic(pdf_path):
    ann = []
    try:
        with fi.open(pdf_path) as doc:
            for page in range(len(doc)):
                pn = doc[page]
                pannot = pn.annots()
                if pannot:
                    for annotation in pannot:
                        if annotation.type[0] == fi.PDF_ANNOT_TEXT:
                            annot_text = annotation.info["content"]
                            ann.append(annot_text)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
    return ann

def extract_text_from_blue_boxes(images):
    lower_blue = np.array([100, 100, 100])
    upper_blue = np.array([130, 255, 255])

    executors = []
    reviewers = []

    for image in images:
        open_cv_image = np.array(image)
        image_bgr = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2BGR)
        hsv_image = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2HSV)
        mask = cv2.inRange(hsv_image, lower_blue, upper_blue)
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        executor_info, reviewer_info = "NA", "NA" 
        
        if contours:
            main_contour = max(contours, key=cv2.contourArea)
            x, y, w, h = cv2.boundingRect(main_contour)
            img_crop = image_bgr[y:y+h, x:x+w]
            gray_img_crop = cv2.cvtColor(img_crop, cv2.COLOR_BGR2GRAY)

            text = pytesseract.image_to_string(gray_img_crop, config='--psm 6')
            lines = text.split('\n')

            for line in lines:
                if 'By' in line:
                    if not executor_info or executor_info == "NA":
                        executor_info = line.split('By', 1)[1].strip()
                    else:
                        reviewer_info = line.split('By', 1)[1].strip()

        executors.append(executor_info)
        reviewers.append(reviewer_info)

    return executors, reviewers

def extract_red_text(images):
    extracted_text = []
    
    lower_red1 = np.array([0, 50, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 50, 50])
    upper_red2 = np.array([180, 255, 255])

    for image in images:
        open_cv_image = np.array(image)
        image_bgr = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2BGR)
        hsv_image = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2HSV)

        mask1 = cv2.inRange(hsv_image, lower_red1, upper_red1)
        mask2 = cv2.inRange(hsv_image, lower_red2, upper_red2)
        mask = mask1 | mask2

        red_areas = cv2.bitwise_and(image_bgr, image_bgr, mask=mask)
        gray_red_areas = cv2.cvtColor(red_areas, cv2.COLOR_BGR2GRAY)
        text = pytesseract.image_to_string(gray_red_areas, config='--psm 6')
        sentences = text.strip().split('\n')
        sentences = [sentence for sentence in sentences if sentence.strip()]
        
        extracted_text.extend(sentences)
    
    return extracted_text

def extract_designation_numbers(images):
    design_numbers = []

    for image in images:
        open_cv_image = np.array(image)
        image_bgr = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2BGR)
        h, w, _ = image_bgr.shape

        roi = image_bgr[int(h*0.7):h, int(w*0.5):w]
        gray_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        gray_roi = cv2.medianBlur(gray_roi, 3)

        _, thresh_roi = cv2.threshold(gray_roi, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(thresh_roi, config=custom_config)

        designation_number = "NA"
        for word in text.split():
            if word.isdigit() and len(word) == 10:
                designation_number = word
                break

        design_numbers.append(designation_number)

    return design_numbers

def preprocess_text(text):
    words = text.split()
    corrected_words = []

    for word in words:
        spell = SpellChecker()
        corrected_word = spell.correction(word) 
        if corrected_word: 
            corrected_words.append(corrected_word)
        else:
            corrected_words.append(word) 
    corrected_words = [word for word in corrected_words if word is not None]

    processed_sentence = ' '.join(corrected_words)

    return processed_sentence

def extract_name_and_date(text):
    date_pattern = r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* \d{1,2}, \d{4}\b"
    date_match = re.search(date_pattern, text)
    date = date_match.group(0) if date_match else "NA"

    text = text.split(' at ')[0].strip()

    name = re.sub(r'[^A-Za-z ]', '', text).strip()

    if not name:
        name = "NA"

    return name, date

def process_pdfs(pdf_paths, output_excel='output.xlsx', contains_handwritten=False):
    
    if os.path.exists(output_excel):
        workbook = openpyxl.load_workbook(output_excel)
        sheet = workbook.active
        row = sheet.max_row + 1  
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Extracted Information"

        sheet["A1"] = "Sr no."
        sheet["B1"] = "Design Number"
        sheet["C1"] = "Executor"
        sheet["D1"] = "Executed Date" 
        sheet["E1"] = "Reviewer"
        sheet["F1"] = "Reviewed Date"  
        sheet["G1"] = "Comments"
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 40
        row = 2 

    for pdf_path in pdf_paths:
        images = convert_from_path(pdf_path)
        if images:
            design_numbers = extract_designation_numbers(images)
            executors, reviewers = extract_text_from_blue_boxes(images)
            red_texts = extract_red_text(images)
            handwritten_texts = handwritten(images)
            iconic_texts = iconic(pdf_path)

            comments = []
            comments.extend(iconic_texts)
            comments.extend(red_texts)
            comments.extend(handwritten_texts) 

            comments_str = "\n".join([preprocess_text(comment) for comment in comments])

            for design_number, executor, reviewer in zip(design_numbers, executors, reviewers):
                executor_name, executor_date = extract_name_and_date(executor)
                reviewer_name, reviewer_date = extract_name_and_date(reviewer)
                
                sheet[f"A{row}"] = row - 1
                sheet[f"B{row}"] = design_number
                sheet[f"C{row}"] = executor_name
                sheet[f"D{row}"] = executor_date
                sheet[f"E{row}"] = reviewer_name
                sheet[f"F{row}"] = reviewer_date
                sheet[f"G{row}"] = comments_str
                
                sheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                sheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                sheet[f"C{row}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                sheet[f"D{row}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                sheet[f"E{row}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                sheet[f"F{row}"].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                sheet[f"G{row}"].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

                line_count = comments_str.count('\n') + 1
                height_per_line = 15
                total_height = line_count * height_per_line
                sheet.row_dimensions[row].height = total_height

                row += 1

    workbook.save(output_excel)

    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        max_length = 0
        for cell in sheet[col]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col].width = max_length + 2 

    messagebox.showinfo("Success", f"Information extracted and saved to {output_excel}")

def select_files(handwritten_option_value):
    files = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    if files:
        process_pdfs(files, contains_handwritten=handwritten_option_value == 1)

def create_gui():
    screen = Tk()
    screen.geometry("400x500+600+50")
    screen.title("Pic2Sheet")
    screen.configure(background="#0a3542") 

    handwritten_option = IntVar(value=0)

    Label(text="Pic2Sheet", font=("Arial", 30), bg="#0a3542", fg="white").pack(pady=50) 

    Button(text="Process PDFs", font=("Arial", 20), height=2, width=20, bg="#5c909c", fg="white", bd=0, command=lambda: select_files(handwritten_option.get())).pack(pady=20)
    Checkbutton(screen, text="Contains handwritten notes", variable=handwritten_option, onvalue=1, offvalue=0, font=("Arial", 12), bg="#0a3542", fg="white", selectcolor="#5c909c").pack(pady=10)


    screen.mainloop()

if __name__ == "__main__":
    create_gui()